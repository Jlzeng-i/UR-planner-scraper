import re
import pandas as pd
from collections import defaultdict
import openpyxl
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import gspread
from gspread_formatting import *
from gspread_formatting.dataframe import format_with_dataframe
from oauth2client.service_account import ServiceAccountCredentials

max_values = {
    "Element Damage Dealt": 116.64,
    "Hit Rate": 58.52,
    "Max Ammunition Capacity": 341.48,
    "ATK": 58.52,
    "Charge Damage": 58.52,
    "Charge Speed": 24.36,
    "Critical Rate": 28.28,
    "Critical Damage": 81.44,
    "DEF": 58.52
}

#Have to setup google cloud in order to use this
scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('credentials.json', scope)
gc = gspread.authorize(creds)

#Change this if you want to run your own sheet
spreadsheet = gc.open_by_url("https://docs.google.com/spreadsheets/d/11t3Uog-D0GaH165NEytUa9ZtiniLWIUZQhLPrkS4AP0/edit?pli=1&gid=0#gid=0")

def percent_to_float(percent_str):
    try:
        return float(percent_str.strip('%'))
    except Exception:
        return 0.0

def convert_to_int(val):
    try:
        if isinstance(val, str) and val.strip() == "":
            return ""
        return int(val)
    except:
        return ""

def make_sheet(filename, output_file):
    with open(filename, "r") as f:
        text = f.read()

    pattern = re.compile(
        r'(?P<prefix>[^\s]+)\s+(?P<name>.+?)\s*(?P<effects>Equipment Effects.*?(\n|$)|\|.*?(\n|$)|$)',
        re.DOTALL
    )

    all_effect_names = set()
    temp_grouped_data = defaultdict(list)

    standard_effect_order = [
        "Element Damage Dealt",
        "ATK",
        "Max Ammunition Capacity",
        "Critical Rate",
        "Critical Damage",
        "Charge Speed",
        "Charge Damage",
        "Hit Rate",
        "DEF"
    ]

    additional_fields = {"Skill 1", "Skill 2", "Skill 3", "Level", "Rarity", "Phase"}

    for match in pattern.finditer(text):
        prefix = match.group("prefix")
        name = match.group("name").strip()
        effects_raw = match.group("effects") or ""
    
        # Initialize effect_dict
        effect_dict = {"Name": name}
    
        # Split effects and extra info
        if effects_raw.startswith("Equipment Effects"):
            effects_part = effects_raw[len("Equipment Effects"):].split("|")[0]
            extra_part = effects_raw[len("Equipment Effects") + len(effects_part):]
        else:
            effects_part = ""
            extra_part = effects_raw
    
        # Parse effect values
        effects = re.findall(r'(Increase [A-Za-z ]+?)\s*([0-9.]+%)', effects_part)
        for eff_name, value in effects:
            clean_name = eff_name.strip()[9:]  # Remove "Increase "
            clean_name = re.sub(r'\s+', ' ', clean_name)
            effect_dict[clean_name] = value
            all_effect_names.add(clean_name)
    
        # Parse Skills
        for skill_num, skill_val in re.findall(r'Skill (\d):\s*(\d+)', extra_part):
            effect_dict[f"Skill {skill_num}"] = skill_val
    
        # Parse Level
        level_match = re.search(r"\['LV(\d+)'\]", extra_part)
        if level_match:
            effect_dict["Level"] = level_match.group(1)
    
        # Parse Rarity (R, SR, SSR)
        rarity_match = re.search(r'\b(R|SR|SSR)\b', extra_part)
        effect_dict["Doll"] = rarity_match.group(1) if rarity_match else ""
    
        # Parse Phase (0-15)
        phase_match = re.search(r'Phase (\d+)', extra_part)
        effect_dict["Phase"] = phase_match.group(1) if phase_match else ""
    
        temp_grouped_data[prefix].append(effect_dict)

    # Build full entries
    for prefix, entries in temp_grouped_data.items():
        for entry in entries:
            full_entry = {"Name": entry["Name"]}
            for effect in all_effect_names:
                full_entry[effect] = entry.get(effect, "0%")
            for field in additional_fields:
                full_entry[field] = entry.get(field, "")
            #grouped_data[prefix].append(full_entry)


    #####FOR GOOGLE SHEETS BUT I GET RATE LIMITED 
    # for prefix, entries in temp_grouped_data.items():
        # base_columns = ["Name", "Level", "Skill 1", "Skill 2", "Skill 3"]
        # df = pd.DataFrame(entries)
        # for col in base_columns:
            # if col not in df.columns:
                # df[col] = "0"
            # if col != "Name":
                # df[col] = df[col] = df[col].apply(convert_to_int)
        # 
        # for eff in standard_effect_order:
            # if eff not in df.columns:
                # df[eff] = "0%"
            # df[eff] = df[eff].apply(lambda x: percent_to_float(x) if isinstance(x, str) else x)
        # 
# 
        # column_order = base_columns + standard_effect_order
        # df = df[column_order]
        # df = df.fillna(0)
        # try:
            # worksheet = spreadsheet.worksheet(prefix)
        # except:
            # worksheet = spreadsheet.add_worksheet(title=prefix[:31], rows="100", cols="20")
        # 
        # rules = get_conditional_format_rules(worksheet)
        # worksheet.clear()
        # rules.clear()
        # worksheet.update([df.columns.values.tolist()] + df.values.tolist())
        # 
        ##Skill rule
        # rule = ConditionalFormatRule(
            # ranges=[GridRange.from_a1_range('C1:E100', worksheet)],
            # gradientRule=GradientRule(
                # maxpoint=InterpolationPoint(color=Color(255/255, 133/255, 222/255), type='NUMBER', value="10"),
                # minpoint=InterpolationPoint(color=Color(1, 1, 1), type='NUMBER', value="0")
            # )
        # )
        # rules.append(rule)
        ##Equip rolls rules
        # start_char= ord('F')
        # for eff in max_values:
            # col_letter = chr(start_char)  # Convert ASCII to column letter
            # cell_range = f'{col_letter}1:{col_letter}100'
            # rule = ConditionalFormatRule(
                # ranges=[GridRange.from_a1_range(cell_range, worksheet)],
                # gradientRule=GradientRule(
                    # maxpoint=InterpolationPoint(color=Color(92/255, 255/255, 255/255), type='NUMBER', value=str(max_values[eff])),
                    # minpoint=InterpolationPoint(color=Color(255/255, 255/255, 255/255), type='NUMBER', value="0")
                # )
            # )
            # rules.append(rule)
            # start_char += 1
        # rules.save()

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        #Define the desired column order
        base_columns = ["Name", "Level", "Skill 1", "Skill 2", "Skill 3", "Doll", "Phase"]
        
        for prefix, entries in temp_grouped_data.items():
            df = pd.DataFrame(entries)
            for col in base_columns:
                if col not in df.columns:
                    df[col] = ""
                if col != "Name" and col != "Doll":
                    df[col] = df[col] = df[col].apply(convert_to_int)
    
            for eff in standard_effect_order:
                if eff not in df.columns:
                    df[eff] = "0%"
                df[eff] = df[eff].apply(lambda x: percent_to_float(x) if isinstance(x, str) else x)

            column_order = base_columns + standard_effect_order
            df = df[column_order]

            sheet_name = prefix[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            worksheet = writer.sheets[sheet_name]
            for col_idx, column in enumerate(df.columns, 1):
                max_length = max(
                    df[column].astype(str).map(len).max(),
                    len(column)
                )
                adjusted_width = max_length + 2 
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

#THIS IS FOR EXCEL
def apply_formatting(in_sheet, out_sheet):
    wb = load_workbook(in_sheet)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        for col_idx, header in enumerate(headers, start=1):
            if header in max_values:
                col_letter = get_column_letter(col_idx)
                max_val = max_values[header]

                color_rule = ColorScaleRule(
                    start_type='num', start_value=0, start_color='FFFFFF',
                    end_type='num', end_value=max_val, end_color='006400'
                )
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    color_rule
                )
    wb.save(out_sheet)


if __name__ == "__main__":
    make_sheet("data.txt", "equipment_effects.xlsx")
    #apply_formatting("equipment_effects.xlsx", "equipment_effects_final.xlsx")
